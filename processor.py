import os
import json
import io
import datetime
import logging
import pdfplumber
import openpyxl
from openai import OpenAI

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

client = OpenAI(
    api_key=os.environ["DEEPSEEK_API_KEY"],
    base_url="https://api.deepseek.com"
)

CRITICAL_FIELDS = [
    "carrier_name", "load_id", "pickup_date", "delivery_date",
    "origin_city", "origin_state", "destination_city", "destination_state",
    "equipment_type", "total_rate"
]

EXTRACTION_PROMPT = (
    "You are a freight document extraction assistant. "
    "From the provided document text, extract the following fields. "
    "Return a JSON object exactly like: "
    "{\"fields\": {\"carrier_name\": {\"value\": \"...\", \"confidence\": 0.0}, ...}, \"overall_confidence\": 0.0}. "
    "Fields to extract: carrier_name, carrier_mc, customer_name, load_id, carrier_pro, customer_po, "
    "pickup_date, delivery_date, origin_city, origin_state, origin_zip, destination_city, destination_state, "
    "destination_zip, equipment_type, commodity, weight, pieces, total_rate, rate_type, miles, fsc, "
    "detention_rate, lumper_fee, tonu, accessorials, payment_terms, document_type. "
    "If a field is not found, set value to null and confidence to 0.0. Dates in ISO-8601 (YYYY-MM-DD)."
)

def extract_fields(text: str) -> dict:
    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=[
            {"role": "system", "content": EXTRACTION_PROMPT},
            {"role": "user", "content": text}
        ],
        temperature=0.0,
        max_tokens=2000
    )
    content = response.choices[0].message.content
    try:
        clean = content.strip().replace("'```json", "").replace("```", "").strip()
        result = json.loads(clean)
        return result
    except json.JSONDecodeError:
        logger.error("Failed to parse AI response as JSON: %s", content)
        return {"fields": {}, "overall_confidence": 0.0}

def determine_status(fields: dict) -> str:
    # Check critical fields missing
    for field in CRITICAL_FIELDS:
        f = fields.get(field, {})
        value = f.get("value") if isinstance(f, dict) else None
        if value is None or value == "":
            return "Missing:critical"
    # Check critical fields low confidence
    for field in CRITICAL_FIELDS:
        f = fields.get(field, {})
        conf = f.get("confidence", 0.0) if isinstance(f, dict) else 0.0
        if conf < 0.7:
            return "Low Confidence:warning"
    # Check non-critical fields missing or low confidence -> Needs Review:warning
    for field_name, f in fields.items():
        if field_name in CRITICAL_FIELDS:
            continue
        if isinstance(f, dict):
            val = f.get("value")
            conf = f.get("confidence", 0.0)
            if val is None or val == "" or conf < 0.7:
                return "Needs Review:warning"
    return "Valid:good"

def process_file(file_bytes: bytes) -> list[dict]:
    raw_text = ""
    # Try PDF
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            pages = [page.extract_text() or "" for page in pdf.pages]
        raw_text = "\n".join(pages)
        if raw_text.strip():
            logger.info("PDF parsed successfully")
    except Exception as e:
        logger.info("PDF parsing failed: %s", e)

    # Try Excel if no text
    if not raw_text.strip():
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            rows = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    rows.append(",".join(str(c) if c is not None else "" for c in row))
            raw_text = "\n".join(rows)
            if raw_text.strip():
                logger.info("Excel parsed successfully")
        except Exception as e:
            logger.info("Excel parsing failed: %s", e)

    # Fallback to plain text
    if not raw_text.strip():
        raw_text = file_bytes.decode("utf-8", errors="ignore")
        logger.info("Using plain text fallback")

    extracted = extract_fields(raw_text)
    fields = extracted.get("fields", {})

    # Title = carrier_name
    carrier = fields.get("carrier_name", {})
    title = carrier.get("value") if isinstance(carrier, dict) else None
    if not title:
        title = "Unknown Carrier"

    # due_date = pickup_date or delivery_date
    pickup = fields.get("pickup_date", {})
    pickup_val = pickup.get("value") if isinstance(pickup, dict) else None
    delivery = fields.get("delivery_date", {})
    delivery_val = delivery.get("value") if isinstance(delivery, dict) else None
    due_date = pickup_val or delivery_val or None
    if due_date and isinstance(due_date, str):
        try:
            # Ensure ISO-8601
            datetime.date.fromisoformat(due_date)
        except ValueError:
            due_date = None

    # details: all fields except carrier_name
    details = {}
    for k, v in fields.items():
        if k != "carrier_name":
            if isinstance(v, dict):
                details[k] = v.get("value")
            else:
                details[k] = v

    status = determine_status(fields)

    return [{
        "title": title,
        "status": status,
        "details": details,
        "due_date": due_date
    }]
